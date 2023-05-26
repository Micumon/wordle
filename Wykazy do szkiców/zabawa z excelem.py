from openpyxl import *

path = "C:\\Users\\admin\\Desktop\\excel\\obręb 24 Podbuszyce - Protokół ustalenia.xlsx"
path_name_pos_plot = "C:\\Users\\admin\\Desktop\\excel\\włas pos dz.xlsx"
path_plot_pos_own = "C:\\Users\\admin\\Desktop\\excel\\dz pos wlas.xlsx"


wb_obj = load_workbook(path)

sheet = wb_obj.active


def sortuj_po_liczbie(element):
    liczba = ""
    for l in element:
        if l.isdigit():
            liczba += l
        elif l == "/":
            liczba += "."
    return float(liczba)


def string_convert_sort(tup):
    tup = list(tup)
    tup.sort(key=sortuj_po_liczbie)
    return ", ".join(tup)


def pers_pos_plot(sheet):
    result = {}
    rows = sheet.max_row
    for row in range(rows):
        row += 1
        owner = sheet.cell(row=row, column=3).value
        if owner is not None and owner != "Podmioty uprawnione do udziału w czynnościach" and owner != "3":
            owner = owner.split("\n")
        else:
            continue
        position = sheet.cell(row=row, column=1).value if sheet.cell(row=row, column=1).value != "" else \
            sheet.cell(row=row - 1, column=1).value
        for person in owner:
            if person in result.keys():
                result[person][1] += (sheet.cell(row=row, column=2).value.replace("\n", " "),) \
                    if sheet.cell(row=row, column=2).value.replace("\n", " ") not in result[person][1] else ()
                result[person][0] += (position,) if position not in result[person][0] else ()
            else:
                result.update({person: [(position,), (sheet.cell(row=row, column=2).value.replace("\n", " "),)]})
    return result


def plot_pos_pers(sheet):
    result = []
    rows = sheet.max_row
    for row in range(rows):
        row += 1
        plot = sheet.cell(row=row, column=2).value
        position = sheet.cell(row=row, column=1).value if sheet.cell(row=row, column=1).value != "" else \
            sheet.cell(row=row - 1, column=1).value
        owner = sheet.cell(row=row, column=3).value
        if plot is not None and position is not None and owner is not None and plot != "2" and plot != \
                "Nr działek ew. do których należy ustalona granica" and plot != "":
            plot = plot.replace("\n", " ")
            result.append([plot, position, owner])
    return result


names_positions_plots = pers_pos_plot(sheet)
names_positions_plots_exel = Workbook()
names_positions_plots_exel_sheet = names_positions_plots_exel.active
row_counter = 1
for owner in names_positions_plots:
    names_positions_plots_exel_sheet.cell(row=row_counter, column=1).value = owner
    names_positions_plots_exel_sheet.cell(row=row_counter, column=2).value = \
        string_convert_sort(names_positions_plots[owner][0])
    names_positions_plots_exel_sheet.cell(row=row_counter, column=3).value = \
        string_convert_sort(names_positions_plots[owner][1])
    row_counter += 1
names_positions_plots_exel.save(path_name_pos_plot)


plots_position_owner = plot_pos_pers(sheet)
plots_position_owner_exel = Workbook()
plots_position_owner_exel_sheet = plots_position_owner_exel.active
row_counter = 1
for row in plots_position_owner:
    plots_position_owner_exel_sheet.cell(row=row_counter, column=1).value = row[0]
    plots_position_owner_exel_sheet.cell(row=row_counter, column=2).value = row[1]
    plots_position_owner_exel_sheet.cell(row=row_counter, column=3).value = row[2]
    row_counter += 1
plots_position_owner_exel.save(path_plot_pos_own)
